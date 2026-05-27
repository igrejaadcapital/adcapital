# -*- coding: utf-8 -*-
from collections import defaultdict
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
MONEY_FMT = '#,##0.00'
DATE_FMT = 'DD/MM/YYYY'
LINK_MAX_WIDTH = 60


def _comprovante_url(transacao):
    if not transacao.comprovante:
        return ''
    try:
        return transacao.comprovante.url or ''
    except (ValueError, AttributeError):
        return ''


def _comprovante_flags(transacao):
    url = _comprovante_url(transacao)
    return ('Sim' if url else 'Não', url)


def _write_comprovante_link(ws, row, col_url, url):
    cell = ws.cell(row=row, column=col_url, value=url or '')
    if url:
        cell.hyperlink = url
        cell.style = 'Hyperlink'


def _autosize_columns(ws, min_width=10, max_width=50):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(cell.value or '')) for cell in col_cells)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def _write_header_row(ws, headers):
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _append_transaction_rows(ws, queryset):
    for tx in queryset:
        tem_anexo, link = _comprovante_flags(tx)
        ws.append([
            tx.id,
            tx.data,
            'Entrada' if tx.tipo == 'ENTRADA' else 'Saída',
            tx.categoria,
            tx.descricao,
            float(tx.valor),
            tem_anexo,
            link,
        ])
        row = ws.max_row
        ws.cell(row=row, column=2).number_format = DATE_FMT
        ws.cell(row=row, column=6).number_format = MONEY_FMT
        _write_comprovante_link(ws, row, 8, link)


def _sheet_lancamentos(ws, queryset):
    ws.title = 'Lançamentos'
    headers = [
        'ID', 'Data', 'Tipo', 'Categoria', 'Descrição',
        'Valor (R$)', 'Comprovante', 'Link do comprovante',
    ]
    _write_header_row(ws, headers)
    _append_transaction_rows(ws, queryset)
    _autosize_columns(ws, max_width=LINK_MAX_WIDTH)


def _sheet_tipo(ws, queryset, title):
    ws.title = title
    headers = [
        'ID', 'Data', 'Categoria', 'Descrição',
        'Valor (R$)', 'Comprovante', 'Link do comprovante',
    ]
    _write_header_row(ws, headers)
    for tx in queryset:
        tem_anexo, link = _comprovante_flags(tx)
        ws.append([
            tx.id,
            tx.data,
            tx.categoria,
            tx.descricao,
            float(tx.valor),
            tem_anexo,
            link,
        ])
        row = ws.max_row
        ws.cell(row=row, column=2).number_format = DATE_FMT
        ws.cell(row=row, column=5).number_format = MONEY_FMT
        _write_comprovante_link(ws, row, 7, link)
    _autosize_columns(ws, max_width=LINK_MAX_WIDTH)


def _sheet_resumo(ws, queryset, periodo_label=''):
    ws.title = 'Resumo'
    gerado_em = timezone.localtime().strftime('%d/%m/%Y %H:%M')
    ws.append(['Relatório financeiro — Igreja AD Capital'])
    ws.append(['Gerado em', gerado_em])
    if periodo_label:
        ws.append(['Período', periodo_label])
    ws.append([])

    total_entradas = queryset.filter(tipo='ENTRADA').aggregate(t=Sum('valor'))['t'] or Decimal('0')
    total_saidas = queryset.filter(tipo='SAIDA').aggregate(t=Sum('valor'))['t'] or Decimal('0')
    saldo = total_entradas - total_saidas

    start_row = ws.max_row + 1
    ws.append(['Total de entradas (R$)', float(total_entradas)])
    ws.append(['Total de saídas (R$)', float(total_saidas)])
    ws.append(['Saldo do período (R$)', float(saldo)])
    for r in range(start_row, start_row + 3):
        ws.cell(row=r, column=2).number_format = MONEY_FMT
    ws.append([])

    ws.append(['Resumo por categoria'])
    _write_header_row(ws, ['Tipo', 'Categoria', 'Quantidade', 'Total (R$)'])
    por_categoria = defaultdict(lambda: {'qtd': 0, 'total': Decimal('0')})
    for tx in queryset:
        key = (tx.tipo, tx.categoria)
        por_categoria[key]['qtd'] += 1
        por_categoria[key]['total'] += tx.valor
    for (tipo, categoria), dados in sorted(por_categoria.items(), key=lambda x: (x[0][0], x[0][1])):
        tipo_label = 'Entrada' if tipo == 'ENTRADA' else 'Saída'
        ws.append([tipo_label, categoria, dados['qtd'], float(dados['total'])])
        ws.cell(row=ws.max_row, column=4).number_format = MONEY_FMT
    ws.append([])

    ws.append(['Resumo mensal'])
    _write_header_row(ws, ['Ano', 'Mês', 'Entradas (R$)', 'Saídas (R$)', 'Saldo (R$)'])
    mensal = defaultdict(lambda: {'entrada': Decimal('0'), 'saida': Decimal('0')})
    for tx in queryset:
        key = (tx.data.year, tx.data.month)
        if tx.tipo == 'ENTRADA':
            mensal[key]['entrada'] += tx.valor
        else:
            mensal[key]['saida'] += tx.valor
    for (ano, mes), vals in sorted(mensal.items()):
        saldo_mes = vals['entrada'] - vals['saida']
        ws.append([ano, mes, float(vals['entrada']), float(vals['saida']), float(saldo_mes)])
        row = ws.max_row
        for col in (3, 4, 5):
            ws.cell(row=row, column=col).number_format = MONEY_FMT

    _autosize_columns(ws)


def build_contabilidade_workbook(
    queryset,
    *,
    incluir_lancamentos=True,
    incluir_abas_separadas=True,
    incluir_resumo=True,
    periodo_label='',
):
    """Monta workbook Excel para envio à contabilidade."""
    wb = Workbook()
    used_first = False

    def _new_sheet(title):
        nonlocal used_first
        if not used_first:
            ws = wb.active
            used_first = True
        else:
            ws = wb.create_sheet()
        ws.title = title
        return ws

    has_data = queryset.exists()

    if incluir_lancamentos and has_data:
        _sheet_lancamentos(_new_sheet('Lançamentos'), queryset)

    if incluir_abas_separadas and has_data:
        entradas = queryset.filter(tipo='ENTRADA')
        saidas = queryset.filter(tipo='SAIDA')
        if entradas.exists():
            _sheet_tipo(_new_sheet('Entradas'), entradas, 'Entradas')
        if saidas.exists():
            _sheet_tipo(_new_sheet('Saídas'), saidas, 'Saídas')

    if incluir_resumo:
        _sheet_resumo(_new_sheet('Resumo'), queryset, periodo_label)

    if not used_first:
        ws = wb.active
        ws.title = 'Relatório'
        ws.append(['Nenhum lançamento encontrado para os filtros informados.'])

    return wb


def workbook_to_bytes(wb):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
